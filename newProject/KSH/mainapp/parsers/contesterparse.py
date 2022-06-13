import datetime
from itertools import groupby

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from googleforms import googleforms

now = datetime.datetime.now()
url = str(input("Ссылка на турнирную таблицу:"))
grant_places = int(input("Количество грантовых мест: "))
if url.__contains__('contestscoreboard'):
    url = url.replace('contestscoreboard', 'contestscoreboardgrid')
page = requests.get(url)
students_list = list()
new_students_list = list()
done_tasks = list()
rating = list()
now = datetime.datetime.now()
namefile = 'KSH' + str(now.year) + '.xlsx'
soup = BeautifulSoup(page.text, 'lxml')
tables = soup.findAll('table')
students_table = tables[3]
table_body = students_table.find('tbody')
students = table_body.findAll('td', attrs={'width': '25%'})
tds = table_body.findAll('td', attrs={'width': '4%'})
for student in students:
    for td in tds:
        students_list.append(student.text)
        new_students_list = [el for el, _ in groupby(students_list)]
        done_tasks.append(td.text)
new_students_list.remove('Участник')
done_tasks.remove('+')
final_done_tasks = done_tasks[:len(new_students_list)]
index = 1
for index in range(len(final_done_tasks)):
    rating.append(str(index + 1))
workbook = Workbook()
group = Workbook()
worksheet = workbook.active
worksheet["A1"] = "Фамилия"
worksheet["B1"] = "Рейтинг"
worksheet["C1"] = "Количество выполненных задач"
second_name = []
row_stud = 2
for student in new_students_list:
    worksheet[row_stud][0].value = student[9:-3]
    second_name.append(student[9:-3])
    row_stud += 1
row_rait = 2
for item in rating:
    worksheet[row_rait][1].value = item
    row_rait += 1
row_count = 2
for item_task in final_done_tasks:
    worksheet[row_count][2].value = item_task
    row_count += 1
workbook.save(filename=namefile)
df = pd.read_excel(namefile, sheet_name='Sheet', usecols=[0, 1, 2])
second_name_from_form = ""
row = 0
names = []
for item in googleforms.iterrows():
    item = googleforms.iloc[row]['Фамилия']
    names.append(item)
    row += 1
row_subject = 0
subjects = set()
row_days = 0
for item in googleforms.iterrows():
    item = googleforms.iloc[row_subject]['В какую группу вы хотите?']
    subject_list = item.split(',')
    for subject in subject_list:
        subject = subject.strip()
        subjects.add(subject)
    row_subject += 1
final_data = pd.merge(googleforms, df, on='Фамилия', how='left').sort_values(by=['Рейтинг'], ascending=True)
for name in second_name:
    for item in names:
        if name.__contains__(item):
            final_data.to_excel('final_result.xlsx', index=False)
writer = pd.ExcelWriter("groups.xlsx", engine='openpyxl')
final_row = 0
final_set = set()
for item in final_data.iterrows():
    item = final_data.iloc[final_row]['В какую группу вы хотите?']
    if item.__contains__(','):
        subject_list = item.split(',')
        for subject in subject_list:
            item = subject.strip()
            final_set.add(item)
    else:
        final_set.add(item)
    final_row += 1
for subject in final_set:
    for subject, group_df in final_data.groupby('В какую группу вы хотите?'):
        group_df.to_excel(writer, sheet_name=subject, index=False)
workbook = writer.book
for sheet in workbook.sheetnames:
    worksheet = writer.sheets[sheet]
    for rows in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in rows:
            if cell.row <= grant_places + 1:
                cell.fill = PatternFill(start_color="60FF60", fill_type='solid')

writer.save()
